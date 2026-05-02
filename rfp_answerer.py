#!/usr/bin/env python3
"""
Excel RFP を読み込み、AnythingLLM API で回答を取得して Excel に書き戻すバッチ処理。
CLI: python3 rfp_answerer.py
Web: python3 rfp_answerer.py serve
"""

from __future__ import annotations

import json
import logging
import os
import secrets
import sys
import threading
import time
from pathlib import Path
from typing import Callable

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

DEFAULT_BASE_URL = "https://ai-rag.umeyouhei.work/api/v1"
API_ERROR_PLACEHOLDER = "要確認（API接続エラー）"
REQUEST_TIMEOUT_SEC = 60
DELAY_BETWEEN_QUESTIONS_SEC = 2

SCRIPT_DIR = Path(__file__).resolve().parent
LOG_PATH = SCRIPT_DIR / "rfp_error.log"
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"
WEB_UPLOAD_DIR = SCRIPT_DIR / "web_uploads"

ProgressCallback = Callable[[int, int, str], None]


def setup_logging() -> logging.Logger:
    logger = logging.getLogger("rfp_answerer")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    fh = logging.FileHandler(LOG_PATH, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    )
    logger.addHandler(fh)
    return logger


def env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None or raw.strip() == "":
        return default
    try:
        return int(raw)
    except ValueError as e:
        raise ValueError(f"環境変数 {name} は整数である必要があります: {raw!r}") from e


def env_column_letter(name: str, default: str) -> int:
    raw = (os.getenv(name) or default).strip().upper()
    if not raw:
        raise ValueError(f"環境変数 {name} が空です")
    try:
        return column_index_from_string(raw)
    except ValueError as e:
        raise ValueError(f"環境変数 {name} は有効な列文字である必要があります: {raw!r}") from e


def extract_answer_text(data: object) -> str | None:
    if not isinstance(data, dict):
        return None
    if data.get("error"):
        return None
    text = data.get("textResponse")
    if isinstance(text, str) and text.strip():
        return text.strip()
    for key in ("response", "message", "answer"):
        val = data.get(key)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None


def call_anythingllm_chat(
    session: requests.Session,
    base_url: str,
    workspace: str,
    api_key: str,
    message: str,
    logger: logging.Logger,
) -> tuple[bool, str | None]:
    url = f"{base_url.rstrip('/')}/workspace/{workspace}/chat"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    payload = {"message": message, "mode": "query"}
    try:
        resp = session.post(
            url,
            headers=headers,
            json=payload,
            timeout=REQUEST_TIMEOUT_SEC,
        )
    except requests.Timeout:
        logger.exception("AnythingLLM API タイムアウト: %s", url)
        return False, None
    except requests.RequestException as e:
        logger.error("AnythingLLM API 接続エラー: %s", e, exc_info=True)
        return False, None

    try:
        body = resp.json()
    except json.JSONDecodeError:
        body = None
        logger.error(
            "AnythingLLM API 非JSON応答: status=%s body=%s",
            resp.status_code,
            resp.text[:2000],
        )
        return False, None

    if resp.status_code >= 400:
        logger.error(
            "AnythingLLM API HTTPエラー: status=%s body=%s",
            resp.status_code,
            json.dumps(body, ensure_ascii=False) if isinstance(body, dict) else resp.text[:2000],
        )
        return False, None

    answer = extract_answer_text(body)
    if answer is None:
        logger.error(
            "AnythingLLM API 応答に回答テキストがありません: %s",
            json.dumps(body, ensure_ascii=False)[:4000],
        )
        return False, None

    return True, answer


def process_workbook(
    path: Path,
    session: requests.Session,
    base_url: str,
    workspace: str,
    api_key: str,
    question_col: int,
    answer_col: int,
    header_rows: int,
    logger: logging.Logger,
    *,
    output_path: Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> tuple[int, int, int, Path]:
    """
    Returns: (success_count, skipped_empty_count, error_count, saved_output_path)
    """
    wb = load_workbook(path, read_only=False)
    ws = wb.active

    success = 0
    skipped_empty = 0
    errors = 0

    start_row = header_rows + 1
    max_row = ws.max_row or 0
    total_rows = max(0, max_row - start_row + 1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is not None:
        out_path = output_path
    else:
        out_path = OUTPUT_DIR / f"answered_{path.name}"

    if progress_callback:
        progress_callback(0, max(total_rows, 1), "ブックを読み込みました")

    try:
        for row in range(start_row, max_row + 1):
            cur = row - start_row + 1
            if progress_callback:
                progress_callback(
                    cur,
                    max(total_rows, 1),
                    f"行 {row} / {max_row} を処理中…",
                )

            q_cell = ws.cell(row=row, column=question_col)
            raw = q_cell.value
            if raw is None:
                skipped_empty += 1
                continue
            question = str(raw).strip()
            if not question:
                skipped_empty += 1
                continue

            if progress_callback:
                progress_callback(
                    cur,
                    max(total_rows, 1),
                    f"行 {row}: AnythingLLM に送信中…",
                )

            ok, answer = call_anythingllm_chat(
                session, base_url, workspace, api_key, question, logger
            )
            if ok and answer is not None:
                ws.cell(row=row, column=answer_col).value = answer
                success += 1
            else:
                ws.cell(row=row, column=answer_col).value = API_ERROR_PLACEHOLDER
                errors += 1

            time.sleep(DELAY_BETWEEN_QUESTIONS_SEC)

        if progress_callback:
            progress_callback(
                max(total_rows, 1),
                max(total_rows, 1),
                f"保存中: {out_path.name}",
            )
        wb.save(out_path)
        print(f"保存しました: {out_path}")
    finally:
        wb.close()

    return success, skipped_empty, errors, out_path


def _load_env() -> None:
    load_dotenv(SCRIPT_DIR / ".env")
    load_dotenv()


def _read_processing_settings() -> tuple[str, str, str, int, int, int]:
    api_key = os.getenv("ANYTHINGLLM_API_KEY", "").strip()
    workspace = os.getenv("ANYTHINGLLM_WORKSPACE", "").strip()
    base_url = os.getenv("ANYTHINGLLM_BASE_URL", DEFAULT_BASE_URL).strip()
    if not api_key:
        raise ValueError("環境変数 ANYTHINGLLM_API_KEY が設定されていません。")
    if not workspace:
        raise ValueError("環境変数 ANYTHINGLLM_WORKSPACE が設定されていません。")
    question_col = env_column_letter("QUESTION_COLUMN", "B")
    answer_col = env_column_letter("ANSWER_COLUMN", "C")
    header_rows = env_int("HEADER_ROWS", 1)
    if header_rows < 0:
        raise ValueError("HEADER_ROWS は 0 以上である必要があります。")
    return api_key, workspace, base_url, question_col, answer_col, header_rows


# --- Web (Flask): 単一ジョブ想定（localhost での夜間バッチ用） ---

_job_lock = threading.Lock()
_job: dict = {
    "status": "idle",
    "message": "待機中",
    "upload_id": None,
    "original_name": None,
    "saved_path": None,
    "current_row_index": 0,
    "total_rows": 0,
    "download_token": None,
    "output_basename": None,
    "stats": None,
}


def _job_snapshot() -> dict:
    with _job_lock:
        return {
            "status": _job["status"],
            "message": _job["message"],
            "upload_id": _job["upload_id"],
            "original_name": _job["original_name"],
            "current_row_index": _job["current_row_index"],
            "total_rows": _job["total_rows"],
            "download_token": _job["download_token"],
            "stats": _job["stats"],
        }


def _set_job(**kwargs: object) -> None:
    with _job_lock:
        for k, v in kwargs.items():
            _job[k] = v


def _run_web_job(saved_path: Path, output_basename: str) -> None:
    logger = setup_logging()
    http = requests.Session()

    def on_progress(current: int, total: int, detail: str) -> None:
        _set_job(
            status="running",
            current_row_index=current,
            total_rows=total,
            message=detail,
        )

    try:
        api_key, workspace, base_url, q_col, a_col, hdr = _read_processing_settings()
    except ValueError as e:
        _set_job(status="error", message=str(e), download_token=None, output_basename=None)
        return

    out_path = OUTPUT_DIR / output_basename
    try:
        success, skipped, errs, written = process_workbook(
            saved_path,
            http,
            base_url,
            workspace,
            api_key,
            q_col,
            a_col,
            hdr,
            logger,
            output_path=out_path,
            progress_callback=on_progress,
        )
        token = secrets.token_urlsafe(32)
        with _job_lock:
            tr = max(int(_job.get("total_rows") or 0), 1)
        _set_job(
            status="done",
            message=(
                f"完了（成功 {success} / スキップ {skipped} / エラー {errs}）"
            ),
            current_row_index=tr,
            total_rows=tr,
            download_token=token,
            output_basename=written.name,
            stats={"success": success, "skipped": skipped, "errors": errs},
        )
    except Exception as e:
        logger.exception("Web ジョブ処理中に例外")
        _set_job(
            status="error",
            message=f"処理中にエラーが発生しました: {e}",
            download_token=None,
            output_basename=None,
        )


def create_app() -> "Flask":
    from flask import Flask, jsonify, request, send_file, send_from_directory
    from werkzeug.utils import secure_filename

    _load_env()
    app = Flask(__name__)
    app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

    @app.get("/")
    def index():
        html_path = SCRIPT_DIR / "app.html"
        if not html_path.is_file():
            return ("app.html が見つかりません。", 500)
        return send_file(html_path, mimetype="text/html; charset=utf-8")

    @app.post("/api/upload")
    def api_upload():
        with _job_lock:
            if _job["status"] == "running":
                return jsonify(error="処理中は新しいファイルをアップロードできません。"), 409

        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify(error="file が必要です。"), 400
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify(error=".xlsx のみ対応しています。"), 400

        upload_id = secrets.token_urlsafe(16)
        raw = Path(f.filename).name
        safe = secure_filename(raw) or "upload.xlsx"
        if not safe.lower().endswith(".xlsx"):
            safe = f"{safe}.xlsx"

        WEB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        saved = WEB_UPLOAD_DIR / f"{upload_id}__{safe}"
        f.save(saved)

        _set_job(
            status="uploaded",
            message="アップロード済み。処理開始を押してください。",
            upload_id=upload_id,
            original_name=raw,
            saved_path=str(saved),
            current_row_index=0,
            total_rows=0,
            download_token=None,
            output_basename=None,
            stats=None,
        )
        return jsonify(upload_id=upload_id, original_name=raw)

    @app.post("/api/start")
    def api_start():
        data = request.get_json(silent=True) or {}
        uid = (data.get("upload_id") or "").strip()

        with _job_lock:
            if _job["status"] == "running":
                return jsonify(error="すでに処理中です。"), 409
            if _job["status"] == "done":
                return jsonify(error="新しいファイルをアップロードしてから開始してください。"), 400
            if not uid or uid != _job.get("upload_id"):
                return jsonify(error="アップロードが無効です。もう一度アップロードしてください。"), 400
            saved = Path(str(_job["saved_path"]))
            orig = str(_job["original_name"] or "RFP.xlsx")
            base_name = Path(orig).name
            if not base_name.lower().endswith(".xlsx"):
                base_name += ".xlsx"
            out_name = f"answered_{uid[:10]}_{secure_filename(base_name) or 'output.xlsx'}"

        def runner():
            _set_job(
                status="running",
                message="処理を開始しました…",
                current_row_index=0,
                total_rows=0,
                download_token=None,
                output_basename=out_name,
                stats=None,
            )
            _run_web_job(saved, out_name)

        threading.Thread(target=runner, daemon=True).start()
        return jsonify(ok=True, message="バックグラウンドで処理を開始しました。")

    @app.get("/api/progress")
    def api_progress():
        snap = _job_snapshot()
        body = {
            "status": snap["status"],
            "message": snap["message"],
            "current_row_index": snap["current_row_index"],
            "total_rows": snap["total_rows"],
            "stats": snap["stats"],
        }
        if snap["status"] == "done" and snap.get("download_token"):
            body["download_token"] = snap["download_token"]
        return jsonify(body)

    @app.get("/api/download")
    def api_download():
        token = (request.args.get("token") or "").strip()
        with _job_lock:
            if not token or token != _job.get("download_token"):
                return jsonify(error="ダウンロードリンクが無効です。"), 403
            name = _job.get("output_basename")
        if not name or "/" in name or "\\" in name or name.startswith("."):
            return jsonify(error="不正なファイル名です。"), 400
        path = OUTPUT_DIR / name
        if not path.is_file():
            return jsonify(error="ファイルが見つかりません。"), 404
        return send_from_directory(
            OUTPUT_DIR,
            name,
            as_attachment=True,
            download_name=name,
        )

    return app


def run_web_server() -> None:
    app = create_app()
    port = int(os.getenv("PORT", "5000"))
    host = os.getenv("HOST", "127.0.0.1")
    print(f"RFP 回答 Web: http://{host}:{port}/")
    app.run(host=host, port=port, use_reloader=False, threaded=True)


def main() -> int:
    _load_env()

    logger = setup_logging()

    try:
        api_key, workspace, base_url, question_col, answer_col, header_rows = (
            _read_processing_settings()
        )
    except ValueError as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"input/ に .xlsx ファイルがありません: {INPUT_DIR}")
        return 0

    total_success = 0
    total_skipped = 0
    total_errors = 0

    session = requests.Session()

    for fpath in xlsx_files:
        print(f"処理中: {fpath.name}")
        try:
            s, sk, err, _out = process_workbook(
                fpath,
                session,
                base_url,
                workspace,
                api_key,
                question_col,
                answer_col,
                header_rows,
                logger,
            )
            total_success += s
            total_skipped += sk
            total_errors += err
        except Exception:
            logger.exception("ワークブック処理中に未処理例外: %s", fpath)
            print(f"致命的エラー（詳細は {LOG_PATH}）: {fpath.name}", file=sys.stderr)
            return 1

    print("--- 集計 ---")
    print(f"処理件数: {total_success}")
    print(f"スキップ件数: {total_skipped}")
    print(f"エラー件数: {total_errors}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].lower() in ("serve", "server", "--web", "web"):
        run_web_server()
    else:
        raise SystemExit(main())
